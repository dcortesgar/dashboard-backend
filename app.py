from collections import defaultdict
from copy import deepcopy
from io import BytesIO
import unicodedata

from flask import Flask, jsonify, request
from flask_cors import CORS
import openpyxl

app = Flask(__name__)
CORS(app)

DISCIPLINE_GROUPS = [
    {
        "group": "Infraestructura y Obra Civil",
        "disciplines": [
            "Trazo geometrico",
            "Topografia",
            "Geologia",
            "Geofisica",
            "Geotecnia",
            "Geohidrologia",
            "Hidrologia",
            "Hidrologia y Drenaje",
            "Terracerias",
            "Proyecto Geometrico",
            "Plataforma y Via",
            "Estructuras",
            "Arqueologia",
            "Arquitectura",
            "Interferencias y Canalizaciones",
            "Ambiental",
        ],
    },
    {
        "group": "Energia",
        "disciplines": [
            "Electrico",
            "Catenaria",
            "Equipamiento Industrial",
        ],
    },
    {
        "group": "Operaciones",
        "disciplines": [
            "Material Rodante",
            "Operaciones",
        ],
    },
    {
        "group": "Control, Mando y Senalizacion",
        "disciplines": [
            "Control, Mando y Senalizacion",
            "Telecomunicaciones",
        ],
    },
    {
        "group": "Transversales",
        "disciplines": [
            "Calidad",
            "RAMS",
            "CAPEX y OPEX",
            "BIM",
        ],
    },
]

GROUP_HEADER_COLORS = [
    "3B2051",
    "3B3434",
    "493956",
    "665972",
    "8C54BB",
    "776969",
    "A18CB3",
    "6E5E79",
    "B8AEC0",
    "D0C9D5",
    "B28DD2",
]

DISCIPLINE_ABBR = {
    "Trazo geometrico": "TRA",
    "Topografia": "TOP",
    "Geologia": "GEOL",
    "Geofisica": "GEOF",
    "Geotecnia": "GEO",
    "Geohidrologia": "GHI",
    "Hidrologia": "HID",
    "Hidrologia y Drenaje": "HYD",
    "Terracerias": "TER",
    "Proyecto Geometrico": "PGM",
    "Plataforma y Via": "VIA",
    "Estructuras": "EST",
    "Arqueologia": "ARG",
    "Arquitectura": "ARQ",
    "Interferencias y Canalizaciones": "IFC",
    "Ambiental": "AMB",
    "Electrico": "ELE",
    "Catenaria": "CAT",
    "Equipamiento Industrial": "EEM",
    "Material Rodante": "MR",
    "Operaciones": "OPS",
    "Control, Mando y Senalizacion": "CMS",
    "Telecomunicaciones": "TLC",
    "Calidad": "CAL",
    "RAMS": "RAMS",
    "CAPEX y OPEX": "CYO",
    "BIM": "BIM",
}


def normalize_text(value):
    if value is None:
        return ""
    return str(value).replace("\n", " ").replace("\r", " ").strip()


def normalize_key(value):
    text = normalize_text(value).lower()
    return "".join(
        char for char in unicodedata.normalize("NFD", text) if unicodedata.category(char) != "Mn"
    )


def build_discipline_registry():
    lookup = {}
    code_lookup = {}

    for group_index, group_data in enumerate(DISCIPLINE_GROUPS):
        group_name = group_data["group"]
        group_color = GROUP_HEADER_COLORS[group_index % len(GROUP_HEADER_COLORS)]
        for discipline_index, discipline_name in enumerate(group_data["disciplines"]):
            code = DISCIPLINE_ABBR[discipline_name]
            metadata = {
                "name": discipline_name,
                "code": code,
                "group": group_name,
                "groupColor": group_color,
                "groupIndex": group_index,
                "disciplineIndex": discipline_index,
            }
            lookup[normalize_key(discipline_name)] = metadata
            code_lookup[code] = metadata

    return lookup, code_lookup


DISCIPLINE_LOOKUP, DISCIPLINE_CODE_LOOKUP = build_discipline_registry()


def get_contrasting_text_color(hex_color):
    r = int(hex_color[0:2], 16)
    g = int(hex_color[2:4], 16)
    b = int(hex_color[4:6], 16)
    luminance = 0.299 * r + 0.587 * g + 0.114 * b
    return "000000" if luminance > 160 else "FFFFFF"


def find_header(headers, expected_name):
    expected = normalize_key(expected_name)

    for header in headers:
        if normalize_key(header) == expected:
            return str(header)

    return None


def find_header_by_options(headers, options):
    for option in options:
        matched = find_header(headers, option)
        if matched:
            return matched
    return None


def get_value(row_dict, header_name):
    if not header_name:
        return ""
    return normalize_text(row_dict.get(header_name))


def get_cell_value(values, index):
    if 0 <= index < len(values):
        return values[index]
    return None


def make_fallback_code(name, fallback_code):
    cleaned_code = normalize_text(fallback_code).upper()
    if cleaned_code:
        return cleaned_code

    letters = [char for char in normalize_key(name).upper() if char.isalnum()]
    if not letters:
        return "N/A"
    return "".join(letters[:3]).ljust(3, "X")


def resolve_discipline(name, fallback_code=None):
    raw_name = normalize_text(name)
    raw_code = normalize_text(fallback_code).upper()

    if raw_name:
        exact_match = DISCIPLINE_LOOKUP.get(normalize_key(raw_name))
        if exact_match:
            return deepcopy(exact_match)

    if raw_code:
        code_match = DISCIPLINE_CODE_LOOKUP.get(raw_code)
        if code_match:
            return deepcopy(code_match)

    if not raw_name and not raw_code:
        return None

    return {
        "name": raw_name or raw_code,
        "code": make_fallback_code(raw_name or raw_code, raw_code),
        "group": "Sin clasificar",
        "groupColor": "D8D2DC",
        "groupIndex": len(DISCIPLINE_GROUPS),
        "disciplineIndex": 999,
    }


def discipline_sort_key(metadata):
    return (
        metadata["groupIndex"],
        metadata["disciplineIndex"],
        normalize_key(metadata["name"]),
        metadata["code"],
    )


def build_matrix_payload(interface_rows):
    leaders_by_key = {}
    participants_by_key = {}

    for item in interface_rows:
        leader = item["leader"]
        participant = item["participant"]
        leaders_by_key[(leader["code"], normalize_key(leader["name"]))] = leader
        participants_by_key[(participant["code"], normalize_key(participant["name"]))] = participant

    leaders = sorted(leaders_by_key.values(), key=discipline_sort_key)
    participants = sorted(participants_by_key.values(), key=discipline_sort_key)

    grouped_rows = []
    for leader in leaders:
        leader_interfaces = [
            item
            for item in interface_rows
            if item["leader"]["code"] == leader["code"]
            and normalize_key(item["leader"]["name"]) == normalize_key(leader["name"])
        ]
        leader_interfaces.sort(
            key=lambda item: (
                discipline_sort_key(item["participant"]),
                normalize_text(item["code"]),
            )
        )

        for item in leader_interfaces:
            grouped_rows.append(
                {
                    "group": leader["group"],
                    "groupColor": leader["groupColor"],
                    "groupTextColor": get_contrasting_text_color(leader["groupColor"]),
                    "leaderName": leader["name"],
                    "leaderCode": leader["code"],
                    "participantName": item["participant"]["name"],
                    "participantCode": item["participant"]["code"],
                    "interfaceCode": normalize_text(item["code"]),
                    "criticidad": normalize_text(item["criticidad"]),
                }
            )

    return {
        "columns": [
            {
                "name": participant["name"],
                "code": participant["code"],
                "group": participant["group"],
                "groupColor": participant["groupColor"],
                "groupTextColor": get_contrasting_text_color(participant["groupColor"]),
            }
            for participant in participants
        ],
        "rows": grouped_rows,
    }


def select_source_sheet(workbook_formula, workbook_data):
    if "RCI" in workbook_formula.sheetnames:
        sheet_name = "RCI"
        return workbook_formula[sheet_name], workbook_data[sheet_name]
    return workbook_formula.active, workbook_data.active


@app.route("/api/documents/upload", methods=["POST"])
def upload_rci_excel():
    if "file" not in request.files:
        return jsonify({"error": "No se recibio ningun archivo."}), 400

    file = request.files["file"]

    if file.filename == "":
        return jsonify({"error": "El archivo no tiene nombre."}), 400

    if not (file.filename.endswith(".xlsx") or file.filename.endswith(".xlsm")):
        return jsonify({"error": "Solo se permiten archivos Excel."}), 400

    file_bytes = file.read()
    workbook_formula = openpyxl.load_workbook(BytesIO(file_bytes), data_only=False, keep_vba=True)
    workbook_data = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True, keep_vba=True)
    sheet_formula, sheet_data = select_source_sheet(workbook_formula, workbook_data)

    header_row = 7
    data_start_row = 8
    headers = [cell.value for cell in sheet_formula[header_row]]

    if not any(headers):
        return jsonify({"error": f"La fila {header_row} no contiene encabezados validos."}), 400

    criticidad_header = None
    for header in headers:
        if "criticidad" in normalize_key(header):
            criticidad_header = str(header)
            break

    if criticidad_header is None:
        return jsonify({"error": "No se encontro la columna 'CRITICIDAD' en el archivo."}), 400

    codigo_header = find_header_by_options(headers, ["CODIGO INTERFAZ", "Codigo Interfaz"])
    tramo_header = find_header_by_options(headers, ["Tramo"])
    estado_header = find_header_by_options(headers, ["ESTADO", "Estado"])
    sistema1_header = find_header_by_options(headers, ["SISTEMA 1", "Sistema 1"])
    sistema2_header = find_header_by_options(headers, ["SISTEMA 2", "Sistema 2"])
    subsistema_lider_header = find_header_by_options(
        headers,
        ["SUBSISTEMA LIDER", "Subsistema lider"],
    )
    subsistema_participante_header = find_header_by_options(
        headers,
        ["SUBSISTEMA PARTICIPANTE", "Subsistema participante"],
    )
    acronimo_lider_header = find_header_by_options(
        headers,
        ["ACRONIMO 1", "Acronimo 1", "DISCIPLINA LIDER", "Disciplina lider"],
    )
    acronimo_participante_header = find_header_by_options(
        headers,
        ["ACRONIMO 2", "Acronimo 2", "DISCIPLINA PARTICIPANTE", "Disciplina participante"],
    )

    rows = []
    matrix_interface_rows = []
    criticidad_alta = 0
    criticidad_media = 0
    criticidad_baja = 0
    discipline_counts = defaultdict(int)

    for formula_row, data_row in zip(
        sheet_formula.iter_rows(min_row=data_start_row, values_only=True),
        sheet_data.iter_rows(min_row=data_start_row, values_only=True),
    ):
        if all(value is None for value in formula_row):
            continue

        row_dict = {}
        data_row_dict = {}
        for index in range(len(headers)):
            header_name = headers[index] if headers[index] is not None else f"col_{index}"
            row_dict[str(header_name)] = formula_row[index]
            data_row_dict[str(header_name)] = data_row[index]

        criticidad_raw = get_value(row_dict, criticidad_header)
        criticidad_value = normalize_key(criticidad_raw)

        if criticidad_value == "alta":
            criticidad_alta += 1
        elif criticidad_value == "media":
            criticidad_media += 1
        elif criticidad_value == "baja":
            criticidad_baja += 1

        lider_name = get_value(row_dict, subsistema_lider_header) or normalize_text(get_cell_value(formula_row, 6))
        lider_fallback_code = get_value(row_dict, acronimo_lider_header) or normalize_text(get_cell_value(formula_row, 7))
        participante_name = get_value(row_dict, subsistema_participante_header) or normalize_text(
            get_cell_value(formula_row, 9)
        )
        participante_fallback_code = get_value(row_dict, acronimo_participante_header) or normalize_text(
            get_cell_value(formula_row, 10)
        )

        leader = resolve_discipline(lider_name, lider_fallback_code)
        participant = resolve_discipline(participante_name, participante_fallback_code)

        codigo_interfaz = normalize_text(data_row_dict.get(codigo_header))
        if not codigo_interfaz:
            codigo_interfaz = normalize_text(row_dict.get(codigo_header))
        if codigo_interfaz.startswith("="):
            tramo = get_value(row_dict, tramo_header) or normalize_text(get_cell_value(formula_row, 1))
            numero = normalize_text(get_cell_value(formula_row, 0))
            if tramo and numero:
                codigo_interfaz = f"{tramo[:4]}_{numero[:3]}"

        if leader:
            discipline_counts[(leader["code"], leader["name"])] += 1

        row_payload = {
            "No.": get_cell_value(formula_row, 0),
            "Tramo": row_dict.get(tramo_header) if tramo_header else get_cell_value(formula_row, 1),
            "CODIGO INTERFAZ": codigo_interfaz or None,
            "SISTEMA 1": row_dict.get(sistema1_header) if sistema1_header else None,
            "SUBSISTEMA LIDER": lider_name or None,
            "SISTEMA 2": row_dict.get(sistema2_header) if sistema2_header else None,
            "SUBSISTEMA PARTICIPANTE": participante_name or None,
            "CRITICIDAD": criticidad_raw or None,
            "ESTADO": row_dict.get(estado_header) if estado_header else None,
            "DISCIPLINA": leader["code"] if leader else None,
            "DISCIPLINA_NOMBRE": leader["name"] if leader else None,
            "DISCIPLINA_PARTICIPANTE": participant["code"] if participant else None,
            "DISCIPLINA_PARTICIPANTE_NOMBRE": participant["name"] if participant else None,
        }
        rows.append(row_payload)

        if leader and participant:
            matrix_interface_rows.append(
                {
                    "code": codigo_interfaz or "",
                    "criticidad": criticidad_raw,
                    "leader": leader,
                    "participant": participant,
                }
            )

    total_interfaces = len(rows)
    discipline_summary = [
        {"code": code, "name": name, "count": count}
        for (code, name), count in sorted(
            discipline_counts.items(),
            key=lambda item: (
                -(item[1]),
                discipline_sort_key(resolve_discipline(item[0][1], item[0][0])),
            ),
        )
    ]
    matrix_payload = build_matrix_payload(matrix_interface_rows)

    return jsonify(
        {
            "totalInterfaces": total_interfaces,
            "criticidadAlta": criticidad_alta,
            "criticidadMedia": criticidad_media,
            "criticidadBaja": criticidad_baja,
            "distribution": [
                {"name": "Alta", "value": criticidad_alta},
                {"name": "Media", "value": criticidad_media},
                {"name": "Baja", "value": criticidad_baja},
            ],
            "disciplineSummary": discipline_summary,
            "rows": rows,
            "matrix": matrix_payload,
        }
    )


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=10000, debug=True)
